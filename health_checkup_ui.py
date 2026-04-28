import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import os
import traceback
import re
import json
import threading
import sys

# ==========================================
# 🌈 系統常數與設定 (Config)
# ==========================================
class AppConfig:
    VERSION = "v3.0.20260411"
    LANG = "zh" # 預設中文
    # ... (rest of AppConfig stays same)
    # 語言翻譯
    LANG_DATA = {}
    try:
        if os.path.exists('translations_en.json'):
            with open('translations_en.json', 'r', encoding='utf-8') as f:
                LANG_DATA = json.load(f)
    except Exception as e:
        print(f"Error loading translation file: {e}")
    TRANSLATIONS = {
        "zh": {
            "title": "TOWNSSVILLE HEALTH CENTER",
            "subtitle": "Sugar, Spice, and Everything Nice!",
            "guest_name": "貴賓姓名:",
            "gender": "性別:",
            "male": "男性",
            "female": "女性",
            "pkg_adj": "  方案與項目調整 ",
            "pkg_detail": "詳細清單",
            "add_on": "增加項目",
            "deselect": "退選主方案項目",
            "config": "選配項目",
            "reset": "重新設定",
            "generate": "生成 PDF 報告",
            "generate_excel": "生成 Excel 報告",
            "guest_info": " 受檢者資訊 ",
            "lang_switch": "切換語言 (English)",
            "save_config": "儲存配置",
            "select_none": "全不選",
            "select_all": "全選",
            "pkg_config_title": "方案內含項目配置",
            "pkg_config_msg": "請確認方案內含之擇一/醫師評估項目",
            "save_addon": "儲存選擇",
            "save_deselect": "儲存退選配置",
            "h_select": "選取",
            "h_cat": "類別",
            "h_item": "項目名稱",
            "h_desc": "檢查內容說明",
            "h_price": "價格",
            "addon_title": "增加項目",
            "ex_name": "受檢者姓名",
            "ex_gender": "性別",
            "ex_pkg_name": "主方案名稱",
            "ex_pkg_price": "主方案價格",
            "ex_addon_sec": "[單項加選項目]",
            "ex_addon_sum": "加選項目總計",
            "ex_deduct_sec": "[主方案退選扣除]",
            "ex_deduct_sum": "退選扣除總計",
            "ex_total": "最終預估總金額",
        },
        "en": {
            "title": "TOWNSSVILLE HEALTH CENTER",
            "subtitle": "Sugar, Spice, and Everything Nice!",
            "guest_name": "Name:",
            "gender": "Gender:",
            "male": "Male",
            "female": "Female",
            "pkg_adj": " Package & Item Adjustment ",
            "pkg_detail": "Details",
            "add_on": "Add Items",
            "deselect": "Remove Items",
            "config": "Option Config",
            "reset": "Reset",
            "generate": "Generate PDF Report",
            "guest_info": " Guest Information ",
            "lang_switch": "Switch Language (中文)",
            "save_config": "Save Config",
            "select_none": "None",
            "select_all": "All",
            "pkg_config_title": "Package Option Config",
            "pkg_config_msg": "Please confirm selected options for this package",
            "save_addon": "Save Selection",
            "save_deselect": "Save Deselection",
            "h_select": "Select",
            "h_cat": "Category",
            "h_item": "Checkup Item",
            "h_desc": "Description",
            "h_price": "Price",
            "addon_title": "Add-on Items",
            "ex_name": "Guest Name",
            "ex_gender": "Gender",
            "ex_pkg_name": "Main Package Name",
            "ex_pkg_price": "Main Package Price",
            "ex_addon_sec": "[Add-on Items Detail]",
            "ex_addon_sum": "Add-on Items Total",
            "ex_deduct_sec": "[Removed Items Deduction]",
            "ex_deduct_sum": "Total Deduction",
            "ex_total": "Final Estimated Total",
        }
    }
    
    # 資料類別翻譯 (Category Translations)
    CAT_TRANS = {
        "零輻射高階 3T磁振造影": "High-End 3T MRI (Non-Radiative)",
        "高階雙源電腦斷層": "Dual-Source CT Scan",
        "X光": "X-Ray",
        "超音波": "Ultrasound",
        "腸胃科": "Gastroenterology",
        "心臟科": "Cardiology",
        "眼科": "Ophthalmology",
        "耳鼻喉科": "ENT (Otolaryngology)",
        "血液常規檢查": "Complete Blood Count (CBC)",
        "糖尿病篩檢": "Diabetes Screening",
        "血脂肪檢查": "Lipid Profile",
        "肝膽功能指標": "Liver Function Tests",
        "腎臟功能指標": "Renal Function Tests",
        "血液腫瘤標記": "Tumor Markers",
        "尿液檢查": "Urinalysis",
        "醫師諮詢": "Physician Consultation",
        "服裝": "Apparel",
        "餐點": "Meals",
        "報告": "Report",
        "交通": "Transportation"
    }
    # 預設自動勾選規則 (Auto-presets)
    AUTO_PRESETS = {
        "菁英": ["PSA", "CA-125", "腫瘤", "CEA", "AFP"],
        "鼎級": ["PSA", "CA-125", "腫瘤", "CEA", "AFP", "MRI", "CT"],
        "全方位": ["PSA", "CA-125", "腫瘤", "CEA", "AFP", "MRI", "CT"],
        "進級": ["PSA", "CA-125", "CEA"],
        "高級": ["PSA", "CA-125", "CEA"],
        "頂級": ["PSA", "CA-125", "CEA", "MRI"],
        "AAS": ["PSA", "CA-125", "眼科", "耳鼻喉", "MRI", "CT"],
        "AS": ["PSA", "CA-125", "MRI"],
        "旗艦": ["PSA", "CA-125", "腫瘤", "CEA", "AFP", "MRI", "CT"],
        "尊爵": ["PSA", "CA-125", "腫瘤", "CEA", "AFP", "MRI", "CT"],
        "御璽": ["PSA", "CA-125", "腫瘤", "CEA", "AFP", "MRI", "CT"],
        "心肺": ["心臟", "肺部", "CT"],
        "腸胃": ["胃鏡", "大腸鏡", "腸胃"],
        "腦部": ["腦部", "MRI"]
    }
    # 同義詞重複過濾關鍵字 (中英文對應)
    SYNONYM_FILTER = {
        "主動脈": ["高階雙源", "雙源電腦斷層", "心臟電腦斷層", "CT", "Aorta"],
        "脊椎": ["頸椎", "腰椎", "全脊椎", "磁振造影", "MRI", "Spine"],
        "顯影劑": ["顯影劑", "Contrast"],
        "腸胃鏡": ["胃鏡", "大腸鏡", "Gastro", "Colonoscopy"],
        "Aorta": ["主動脈", "高階雙源", "CT"],
        "Spine": ["脊椎", "MRI", "Spine"],
        "MRI": ["磁振造影"],
        "CT": ["電腦斷層"]
    }
    COLORS = {
        "BG_PINK": "#FFE4E1",
        "HEADER_PINK": "#FF69B4",
        "SEC_BLUE": "#E0F2F1",
        "BTN_BLUE": "#4FC3F7",
        "SEC_GREEN": "#F1F8E9",
        "BTN_GREEN": "#81C784",
        "FOOTER_BLOSSOM": "#FF1493",
        "FOOTER_BUBBLES": "#00BFFF",
        "FOOTER_BUTTERCUP": "#32CD32"
    }
    UI_FONTS = {
        "TITLE": ("微软雅黑", 22, "bold"),
        "BOLD": ("微软雅黑", 14, "bold"),
        "NORMAL": ("微软雅黑", 14)
    }
    CATEGORY_ORDER = [
        "零輻射高階 3T磁振造影", "輔助高階3T 磁振造影影像診斷", "高階雙源電腦斷層", "X光", "超音波",
        "骨質密度", "腸胃科", "心臟科", "眼科", "耳鼻喉科", "婦科(限女性)", "生理量測",
        "心血管健康指標", "血液常規檢查", "糖尿病篩檢", "血脂肪檢查", "電解質", "肝膽功能指標",
        "腎臟功能指標", "胰臟指標", "甲狀腺功能指標", "血液腫瘤標記", "肺癌標記", "凝血功能",
        "栓塞指數", "肝炎篩檢", "類風濕性關節炎篩檢", "過敏原檢查", "荷爾蒙及特殊檢驗",
        "維生素相關", "尿液檢查", "糞便檢查", "醫師諮詢", "餐點", "服裝", "報告", "交通"
    ]
# ==========================================
# 🛠️ 核心處理引擎
# ==========================================
class HealthCheckupApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"🏥 健檢客戶明細助手 {AppConfig.VERSION}")
        self.root.geometry("900x800")
        self.root.configure(bg=AppConfig.COLORS["BG_PINK"])
        
        self.excel_path = '中心全組套.xlsx'
        self.font_path = 'msyh.ttf'
        self.packages = {}
        self.add_on_items = []
        self.selected_addons = []
        self.deselected_main_items = set() 
        self.addon_price_map = {}
        self.pkg_config_selections = {} 
        self.current_raw_pkg = None 
        
        # ⚡ 效能優化快取 (Caching)
        self.data_df = None      
        self.items_df = None     
        self.df_addon_cache = None 
        self._tc_cache = {}      
        self._current_cached_lang = None 
        
        clean_order = [self.clean_cat_name(c) for c in AppConfig.CATEGORY_ORDER]
        self.cat_priority = {name: i for i, name in enumerate(clean_order)}
        self.init_font()
        self.create_widgets()
        self.check_and_load_data()

    def check_and_load_data(self):
        if not os.path.exists(self.excel_path):
            messagebox.showinfo("🔍 找不到資料表", f"在目錄中找不到 '{self.excel_path}'\n請手動選擇正確的 Excel 檔案。")
            f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
            if f: self.excel_path = f
            else: self.root.destroy(); return
        
        if self.load_data(): self.update_combobox()
        else: 
            is_en = AppConfig.LANG == "en"
            msg = "Failed to load Excel data. Please ensure the file is not opened in another program (like Microsoft Excel)." if is_en else "讀取 Excel 內容時發生錯誤。\n\n💡 提示：請確認 Excel 檔案是否已被其他程式（如 Microsoft Excel）開啟。"
            messagebox.showerror("❌ " + ("Load Failed" if is_en else "載入失敗"), msg)

    def init_font(self):
        if os.path.exists(self.font_path):
            try: pdfmetrics.registerFont(TTFont('ChineseFont', self.font_path))
            except: pass
        else: messagebox.showwarning("⚠️ 字型警告", "找不到 msyh.ttf，PDF 中文可能無法正常顯示。")

    @staticmethod
    def safe_parse_price(v):
        if pd.isna(v) or not str(v).strip(): return 0
        try:
            c = re.sub(r'[^\d]', '', str(v))
            return int(c) if c else 0
        except: return 0

    @staticmethod
    def clean_cat_name(n):
        if not n or str(n).lower() == 'nan': return ""
        c = str(n).replace("(", "").replace(")", "").replace(" ", "").replace("　", "").replace("指標", "標記").strip()
        m = {"腸胃": "腸胃科", "婦科": "婦科限女性", "血管健康標記": "心血管健康標記", "類風濕": "類風濕性關節炎篩檢", "荷爾蒙": "荷爾蒙及特殊檢驗"}
        for k, val in m.items():
            if k in c and (k != "血管健康標記" or "心血管健康標記" not in c): return val
        return c

    def load_data(self, force_reload=False):
        try:
            if not force_reload and self.data_df is not None and self._current_cached_lang == AppConfig.LANG:
                return True
            self.packages = {}
            self.add_on_items = []
            self.addon_price_map = {}
            current_path = self.excel_path
            if AppConfig.LANG == "en":
                en_path = self.excel_path.replace(".xlsx", "_EN.xlsx")
                if os.path.exists(en_path): current_path = en_path
            xl = pd.ExcelFile(current_path)
            self.data_df = pd.read_excel(xl, sheet_name=xl.sheet_names[0], header=None)
            self._current_cached_lang = AppConfig.LANG
            df_main = self.data_df
            sc, ec = 3, min(51, len(df_main.columns))
            ns, ps = df_main.iloc[0, sc:ec], df_main.iloc[1, sc:ec]
            last_n = ""
            for i in range(len(ns)):
                n, p = str(ns.iloc[i]).strip(), str(ps.iloc[i]).strip()
                if n == 'nan' or not n: n = last_n
                else: last_n = n
                if n and n != 'nan':
                    g = "男" if "(男)" in p else "女" if "(女)" in p else ""
                    k = f"{n} ({g})" if g else f"{n} (方案{i+1})"
                    self.packages[k] = {"price": p, "col_idx": sc + i}
            self.items_df = df_main.iloc[3:].copy()
            if len(xl.sheet_names) > 1:
                self.df_addon_cache = pd.read_excel(xl, sheet_name=xl.sheet_names[1], header=None)
                df_addon = self.df_addon_cache
                carrying_cat = ""
                for idx in range(1, len(df_addon)):
                    c_val = str(df_addon.iloc[idx, 0]).strip()
                    if c_val and c_val.lower() != 'nan': carrying_cat = c_val
                    item = str(df_addon.iloc[idx, 1]).strip()
                    if item and item.lower() != 'nan':
                        p_str = str(df_addon.iloc[idx, 3]).strip()
                        self.add_on_items.append({"cat": carrying_cat, "item": item, "desc": str(df_addon.iloc[idx, 2]).strip(), "price": p_str, "orig_row": idx})
                        self.addon_price_map[(carrying_cat, item)] = p_str
            return True
        except:
            traceback.print_exc(); return False

    def create_widgets(self):
        self.ui_elements = {}
        header = tk.Frame(self.root, bg=AppConfig.COLORS["HEADER_PINK"], pady=15); header.pack(fill=tk.X)
        self.ui_elements["title"] = tk.Label(header, text=self.gt("title"), font=AppConfig.UI_FONTS["TITLE"], bg=AppConfig.COLORS["HEADER_PINK"], fg="white"); self.ui_elements["title"].pack()
        self.ui_elements["subtitle"] = tk.Label(header, text=self.gt("subtitle"), font=("Comic Sans MS", 10, "italic"), bg=AppConfig.COLORS["HEADER_PINK"], fg="white"); self.ui_elements["subtitle"].pack()
        lang_btn = tk.Button(header, text=self.gt("lang_switch"), command=self.toggle_language, bg=AppConfig.COLORS["BTN_BLUE"], fg="white", font=("微软雅黑", 10, "bold")); lang_btn.place(relx=0.98, rely=0.05, anchor="ne"); self.ui_elements["lang_btn"] = lang_btn
        main_body = tk.Frame(self.root, bg=AppConfig.COLORS["BG_PINK"], padx=40, pady=20); main_body.pack(fill=tk.BOTH, expand=True)
        info_frame = tk.LabelFrame(main_body, text=self.gt("guest_info"), font=AppConfig.UI_FONTS["BOLD"], bg=AppConfig.COLORS["SEC_BLUE"], fg=AppConfig.COLORS["FOOTER_BUBBLES"], padx=20, pady=20, bd=4, relief=tk.RIDGE); info_frame.pack(fill=tk.X, pady=10); self.ui_elements["guest_info_frame"] = info_frame
        lbl_name = tk.Label(info_frame, text=self.gt("guest_name"), font=AppConfig.UI_FONTS["BOLD"], bg=AppConfig.COLORS["SEC_BLUE"]); lbl_name.grid(row=0, column=0, sticky="w"); self.ui_elements["guest_name_lbl"] = lbl_name
        self.name_var = tk.StringVar(); tk.Entry(info_frame, textvariable=self.name_var, width=20, font=AppConfig.UI_FONTS["NORMAL"]).grid(row=0, column=1, padx=10)
        lbl_gender = tk.Label(info_frame, text=self.gt("gender"), font=AppConfig.UI_FONTS["BOLD"], bg=AppConfig.COLORS["SEC_BLUE"]); lbl_gender.grid(row=0, column=2, padx=(30, 10)); self.ui_elements["gender_lbl"] = lbl_gender
        self.gender_var = tk.StringVar(value="男性"); self.rb_male = tk.Radiobutton(info_frame, text=self.gt("male"), variable=self.gender_var, value="男性", font=AppConfig.UI_FONTS["NORMAL"], bg=AppConfig.COLORS["SEC_BLUE"], command=self.update_combobox); self.rb_male.grid(row=0, column=3); self.ui_elements["male_rb"] = self.rb_male
        self.rb_female = tk.Radiobutton(info_frame, text=self.gt("female"), variable=self.gender_var, value="女性", font=AppConfig.UI_FONTS["NORMAL"], bg=AppConfig.COLORS["SEC_BLUE"], command=self.update_combobox); self.rb_female.grid(row=0, column=4); self.ui_elements["female_rb"] = self.rb_female
        pkg_frame = tk.LabelFrame(main_body, text=self.gt("pkg_adj"), font=AppConfig.UI_FONTS["BOLD"], bg="white", fg=AppConfig.COLORS["FOOTER_BLOSSOM"], padx=20, pady=20, bd=4, relief=tk.RIDGE); pkg_frame.pack(fill=tk.X, pady=10); self.ui_elements["pkg_frame"] = pkg_frame
        self.pkg_var = tk.StringVar(); self.pkg_cb = ttk.Combobox(pkg_frame, textvariable=self.pkg_var, width=35, state="readonly", font=AppConfig.UI_FONTS["NORMAL"]); self.pkg_cb.grid(row=0, column=1, padx=10); self.pkg_cb.bind("<<ComboboxSelected>>", lambda e: self.on_main_pkg_selected())
        self.detail_btn = tk.Button(pkg_frame, text=self.gt("pkg_detail"), command=self.open_main_package_window, bg=AppConfig.COLORS["BTN_BLUE"], fg="white", font=AppConfig.UI_FONTS["BOLD"]); self.detail_btn.grid(row=0, column=2, padx=10); self.ui_elements["detail_btn"] = self.detail_btn
        adj_btn_frame = tk.Frame(pkg_frame, bg="white"); adj_btn_frame.grid(row=1, column=0, columnspan=3, pady=(20, 0), sticky="w")
        self.addon_btn = tk.Button(adj_btn_frame, text=self.gt("add_on"), command=self.open_addon_window, state="disabled", bg=AppConfig.COLORS["HEADER_PINK"], fg="white", font=AppConfig.UI_FONTS["BOLD"]); self.addon_btn.pack(side=tk.LEFT, padx=15)
        self.deselect_btn = tk.Button(adj_btn_frame, text=self.gt("deselect"), command=self.open_deselect_window, state="disabled", bg="#FFA07A", fg="white", font=AppConfig.UI_FONTS["BOLD"]); self.deselect_btn.pack(side=tk.LEFT, padx=15)
        self.config_btn = tk.Button(adj_btn_frame, text=self.gt("config"), command=self.open_pkg_config_window, state="disabled", bg="#BA55D3", fg="white", font=AppConfig.UI_FONTS["BOLD"]); self.config_btn.pack(side=tk.LEFT, padx=15)
        reset_btn = tk.Button(adj_btn_frame, text=self.gt("reset"), command=self.reset_fields, bg="#9E9E9E", fg="white", font=AppConfig.UI_FONTS["BOLD"]); reset_btn.pack(side=tk.LEFT, padx=15); self.ui_elements["reset_btn"] = reset_btn
        act_frame = tk.Frame(main_body, bg=AppConfig.COLORS["BG_PINK"]); act_frame.pack(pady=30)
        self.gen_btn = tk.Button(act_frame, text=self.gt("generate"), command=self.generate_report, bg=AppConfig.COLORS["BTN_GREEN"], fg="white", font=AppConfig.UI_FONTS["TITLE"], padx=20, pady=15); self.gen_btn.pack(side=tk.LEFT, padx=10); self.ui_elements["gen_btn"] = self.gen_btn
        self.gen_xlsx_btn = tk.Button(act_frame, text=self.gt("generate_excel"), command=self.generate_excel_report, bg=AppConfig.COLORS["BTN_BLUE"], fg="white", font=AppConfig.UI_FONTS["TITLE"], padx=20, pady=15); self.gen_xlsx_btn.pack(side=tk.LEFT, padx=10); self.ui_elements["gen_xlsx_btn"] = self.gen_xlsx_btn
        footer = tk.Frame(self.root, height=80); footer.pack(fill=tk.X, side=tk.BOTTOM)
        for n, i, c in [("BLOSSOM", "🎀", AppConfig.COLORS["FOOTER_BLOSSOM"]), ("BUBBLES", "🫧", AppConfig.COLORS["FOOTER_BUBBLES"]), ("BUTTERCUP", "⚡", AppConfig.COLORS["FOOTER_BUTTERCUP"])]:
            f = tk.Frame(footer, bg=c); f.pack(side=tk.LEFT, fill="both", expand=True)
            tk.Label(f, text=f"{n} {i}", font=AppConfig.UI_FONTS["BOLD"], bg=c, fg="white").pack(pady=10)

    def gt(self, key): return AppConfig.TRANSLATIONS[AppConfig.LANG].get(key, key)
    def toggle_language(self):
        AppConfig.LANG = "en" if AppConfig.LANG == "zh" else "zh"
        self._tc_cache.clear(); self.load_data(); self.refresh_ui()

    def refresh_ui(self):
        is_en = AppConfig.LANG == "en"; self.root.geometry("900x800")
        self.ui_elements["title"].config(text=self.gt("title")); self.ui_elements["subtitle"].config(text=self.gt("subtitle"))
        self.ui_elements["lang_btn"].config(text=self.gt("lang_switch")); self.ui_elements["guest_info_frame"].config(text=self.gt("guest_info"))
        self.ui_elements["guest_name_lbl"].config(text=self.gt("guest_name")); self.ui_elements["gender_lbl"].config(text=self.gt("gender"))
        self.ui_elements["male_rb"].config(text=self.gt("male")); self.ui_elements["female_rb"].config(text=self.gt("female"))
        self.ui_elements["pkg_frame"].config(text=self.gt("pkg_adj")); self.ui_elements["detail_btn"].config(text=self.gt("pkg_detail"))
        self.ui_elements["reset_btn"].config(text=self.gt("reset")); self.ui_elements["gen_btn"].config(text=self.gt("generate")); self.ui_elements["gen_xlsx_btn"].config(text=self.gt("generate_excel"))
        raw_k = self.current_raw_pkg; self.update_combobox()
        if raw_k: self.pkg_var.set(self.tc(raw_k))
        self.pkg_cb.config(width=50 if is_en else 35); has_pkg = bool(self.pkg_var.get())
        if not self.selected_addons: self.addon_btn.config(text=self.gt("add_on"))
        else: self.addon_btn.config(text=f"{( 'Add-on' if is_en else '已加選')} ({len(self.selected_addons)} {( 'Items' if is_en else '項')})")
        if not self.deselected_main_items: self.deselect_btn.config(text=self.gt("deselect"))
        else: self.deselect_btn.config(text=f"{( 'Remove' if is_en else '已退選')} ({len(self.deselected_main_items)} {( 'Items' if is_en else '項')})")
        cfg_count = sum(1 for v in self.pkg_config_selections.values() if v)
        if cfg_count == 0: self.config_btn.config(text=self.gt("config"))
        else: self.config_btn.config(text=f"{( 'Options' if is_en else '選配項目')} ({cfg_count} {( 'Items' if is_en else '項')})")
        st = "normal" if has_pkg else "disabled"; self.addon_btn.config(state=st); self.deselect_btn.config(state=st); self.config_btn.config(state=st)

    def tc(self, text):
        if AppConfig.LANG == "zh" or not text: return text
        raw_text = str(text).strip()
        if raw_text in self._tc_cache: return self._tc_cache[raw_text]
        clean_text = raw_text.replace("\n", "").replace(" ", ""); result = raw_text
        if raw_text in AppConfig.LANG_DATA: result = AppConfig.LANG_DATA[raw_text]
        else:
            found = False
            for k, v in AppConfig.LANG_DATA.items():
                if k.replace("\n", "").replace(" ", "") == clean_text: result = v; found = True; break
            if not found:
                if raw_text in AppConfig.CAT_TRANS: result = AppConfig.CAT_TRANS[raw_text]
                elif "(" in raw_text and ")" in raw_text:
                    base = raw_text.split("(")[0].strip(); suffix = raw_text.split("(")[1].split(")")[0].strip()
                    trans_base = self.tc(base); m = {"男": "Male", "女": "Female", "限男": "Male Only", "限女": "Female Only"}
                    trans_suffix = m.get(suffix, suffix)
                    if trans_base != base: result = f"{trans_base} ({trans_suffix})"
                else:
                    m = {"男": "Male", "女": "Female", "方案": "Pkg", "限": "Only "}; res = raw_text
                    for k, v in m.items():
                        if k in res: res = res.replace(k, v)
                    result = res
        self._tc_cache[raw_text] = result; return result

    def reset_fields(self):
        self.name_var.set(""); self.gender_var.set("男性"); self.selected_addons = []; self.deselected_main_items = set(); self.pkg_config_selections = {}; self.refresh_ui()
        msg_t = "Reset completed!" if AppConfig.LANG == "en" else "重置完成"; msg_b = "All fields have been reset to default values." if AppConfig.LANG == "en" else "所有欄位已恢復預設值。"
        messagebox.showinfo("🧹 " + msg_t, msg_b)

    def update_combobox(self):
        g = self.gender_var.get(); tag = "(女)" if g == "男性" else "(男)"
        raw_pkgs = [k for k in self.packages.keys() if tag not in k]
        display_pkgs = [self.tc(k) for k in raw_pkgs]
        if display_pkgs:
            self.pkg_cb['values'] = display_pkgs
            if not self.pkg_var.get(): self.pkg_cb.current(0); self.on_main_pkg_selected()

    def on_main_pkg_selected(self):
        display_pkg = self.pkg_var.get()
        if display_pkg:
            pkg_k = None
            for k in self.packages.keys():
                if self.tc(k) == display_pkg: pkg_k = k; break
            if pkg_k:
                if pkg_k != self.current_raw_pkg:
                    self.current_raw_pkg = pkg_k; self.deselected_main_items.clear(); self.pkg_config_selections.clear(); self.apply_auto_presets(pkg_k)
                self.addon_btn.config(state="normal"); self.deselect_btn.config(state="normal"); self.config_btn.config(state="normal")

    def apply_auto_presets(self, pkg_k):
        name_val = self.name_var.get().upper(); col_idx = self.packages[pkg_k]["col_idx"]
        for kw, items in AppConfig.AUTO_PRESETS.items():
            if kw in pkg_k or kw in name_val:
                for idx, row in self.items_df.iterrows():
                    val = str(self.data_df.iloc[idx, col_idx]).strip()
                    if '◎' in val:
                        item_name = str(row.iloc[1]).strip()
                        for target in items:
                            if target.upper() in item_name.upper(): self.pkg_config_selections[item_name] = True

    def open_pkg_config_window(self):
        display_pkg = self.pkg_var.get(); pkg_k = None
        for k in self.packages.keys():
            if self.tc(k) == display_pkg: pkg_k = k; break
        if not pkg_k: return
        col_idx = self.packages[pkg_k]["col_idx"]; g = self.gender_var.get(); config_items = []; tc = ""
        for idx, row in self.items_df.iterrows():
            c_v = str(self.data_df.iloc[idx, 0]).strip()
            if c_v and c_v.lower() != 'nan': tc = c_v
            val = str(row.iloc[col_idx]).strip()
            if '◎' in val:
                item_name = str(row.iloc[1]).strip(); desc = str(row.iloc[2]).strip()
                block_kws = ["卵巢癌", "CA-125", "子宮頸", "乳房", "婦科", "(女)", "限女"] if g == "男性" else ["攝護腺", "PSA", "(男)", "限男"]
                if any(kw in item_name or kw in desc for kw in block_kws): continue
                config_items.append({"cat": tc, "item": item_name, "desc": desc, "idx": idx})
        if not config_items: self.config_btn.config(text="⚙️ 無選配項目", state="disabled"); return
        win = tk.Toplevel(self.root); win.title(f"🦋 {self.gt('pkg_config_title')} 🦋"); win.geometry("900x750"); win.configure(bg=AppConfig.COLORS["SEC_BLUE"]); win.grab_set()
        tk.Label(win, text=f"✨ {self.gt('pkg_config_msg')} ✨", font=AppConfig.UI_FONTS["BOLD"], bg=AppConfig.COLORS["HEADER_PINK"], fg="white", pady=10).pack(fill=tk.X)
        cvs = {it['item']: tk.BooleanVar(value=self.pkg_config_selections.get(it['item'], False)) for it in config_items}
        def save(): self.pkg_config_selections = {it: var.get() for it, var in cvs.items()}; self.refresh_ui(); win.destroy()
        btn_frame = tk.Frame(win, bg=AppConfig.COLORS["SEC_BLUE"]); btn_frame.pack(pady=10)
        tk.Button(btn_frame, text=f"💖 {self.gt('save_config')}", command=save, font=AppConfig.UI_FONTS["TITLE"], bg=AppConfig.COLORS["BTN_GREEN"], fg="white", padx=30).pack(side=tk.LEFT, padx=10)
        def select_all(v):
            for var in cvs.values(): var.set(v)
        tk.Button(btn_frame, text=f"⬜ {self.gt('select_none')}", command=lambda: select_all(False), bg="#9E9E9E", fg="white", font=AppConfig.UI_FONTS["BOLD"]).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text=f"⬛ {self.gt('select_all')}", command=lambda: select_all(True), bg="#757575", fg="white", font=AppConfig.UI_FONTS["BOLD"]).pack(side=tk.LEFT, padx=5)
        cnt = tk.Frame(win, bg="white", bd=3, relief=tk.SUNKEN); cnt.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        can = tk.Canvas(cnt, bg="white"); vsb = ttk.Scrollbar(cnt, orient="vertical", command=can.yview); sf = tk.Frame(can, bg="white")
        sf.bind("<Configure>", lambda e: can.configure(scrollregion=can.bbox("all"))); can.create_window((0,0), window=sf, anchor="nw"); can.configure(yscrollcommand=vsb.set)
        for i, it in enumerate(config_items):
            tk.Checkbutton(sf, variable=cvs[it['item']], font=AppConfig.UI_FONTS["NORMAL"], bg="white").grid(row=i, column=0, sticky="w", padx=10, pady=5)
            disp_txt = f"[{self.tc(it['cat'])}] {self.tc(it['item'])}"; lbl_main = tk.Label(sf, text=disp_txt, font=AppConfig.UI_FONTS["BOLD"], bg="white"); lbl_main.grid(row=i, column=1, sticky="w")
            lbl_desc = tk.Label(sf, text=f"{it['desc']}", font=("微软雅黑", 10), fg="gray", bg="white"); lbl_desc.grid(row=i, column=2, sticky="w", padx=20)
            def toggle_cfg(e, item_key=it['item']): cvs[item_key].set(not cvs[item_key].get())
            lbl_main.bind("<Button-1>", toggle_cfg); lbl_desc.bind("<Button-1>", toggle_cfg)
        can.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")

    def open_deselect_window(self):
        display_pkg = self.pkg_var.get(); pkg_k = None
        for k in self.packages.keys():
            if self.tc(k) == display_pkg: pkg_k = k; break
        if not pkg_k: return
        col_idx = self.packages[pkg_k]["col_idx"]; win = tk.Toplevel(self.root); win.title(f"➖ {self.gt('deselect')} ➖"); win.geometry("800x850"); win.grab_set()
        top = tk.Frame(win, bg="#FFA07A", pady=15); top.pack(fill=tk.X); items = []; tc = ""
        for idx, row in self.items_df.iterrows():
            c_v = str(self.data_df.iloc[idx, 0]).strip()
            if c_v and c_v.lower() != 'nan': tc = c_v
            if str(row.iloc[col_idx]).strip().lower() != 'nan': items.append({"idx": idx, "cat": tc, "item": str(row.iloc[1]).strip()})
        cvs = {it['idx']: tk.BooleanVar(value=(it['idx'] in self.deselected_main_items)) for it in items}
        def save(): self.deselected_main_items = {idx for idx, v in cvs.items() if v.get()}; self.refresh_ui(); win.destroy()
        tk.Button(top, text=f"💾 {self.gt('save_deselect')}", command=save, font=AppConfig.UI_FONTS["BOLD"]).pack(side=tk.LEFT, padx=20)
        cnt = tk.Frame(win, bg="white", bd=3, relief=tk.SUNKEN); cnt.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        can = tk.Canvas(cnt, bg="white"); vsb = ttk.Scrollbar(cnt, orient="vertical", command=can.yview); sf = tk.Frame(can, bg="white")
        sf.bind("<Configure>", lambda e: can.configure(scrollregion=can.bbox("all"))); can.create_window((0,0), window=sf, anchor="nw"); can.configure(yscrollcommand=vsb.set)
        for i, it in enumerate(items):
            row_frame = tk.Frame(sf, bg="white"); row_frame.grid(row=i, column=0, columnspan=2, sticky="nsew")
            cb = tk.Checkbutton(row_frame, variable=cvs[it['idx']], bg="white"); cb.pack(side=tk.LEFT, padx=10)
            disp_txt = f"[{self.tc(it['cat'])}] {self.tc(it['item'])}"; lbl_item = tk.Label(row_frame, text=disp_txt, font=AppConfig.UI_FONTS["BOLD"], bg="white", width=60, anchor="w"); lbl_item.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            def toggle_cb(e, idx=it['idx']): cvs[idx].set(not cvs[idx].get())
            lbl_item.bind("<Button-1>", toggle_cb)
            def on_enter(e, f=row_frame, c=cb, l1=lbl_item):
                for w in [f, c, l1]: w.configure(bg="#FFE4B5")
            def on_leave(e, f=row_frame, c=cb, l1=lbl_item):
                for w in [f, c, l1]: w.configure(bg="white")
            row_frame.bind("<Enter>", on_enter); row_frame.bind("<Leave>", on_leave)
        can.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")
        def _mw_deselect(e): can.yview_scroll(int(-1*(e.delta/120)), "units")
        can.bind_all("<MouseWheel>", _mw_deselect); win.bind("<Destroy>", lambda e: can.unbind_all("<MouseWheel>"))

    def open_addon_window(self):
        display_pkg = self.pkg_var.get(); pkg_k = None
        for k in self.packages.keys():
            if self.tc(k) == display_pkg: pkg_k = k; break
        if not pkg_k: return
        col_idx = self.packages[pkg_k]["col_idx"]; main_items_clean = []
        for idx, row in self.items_df.iterrows():
            val = str(self.data_df.iloc[idx, col_idx]).strip().lower()
            if val != 'nan' and val != "" and idx not in self.deselected_main_items:
                item_name = str(row.iloc[1]).strip().replace("\n", "").replace(" ", "")
                if item_name: main_items_clean.append(item_name)
        filtered = []; g = self.gender_var.get()
        for i, info in enumerate(self.add_on_items):
            ad_item_raw = str(info['item']).strip(); ad_item_clean = ad_item_raw.replace("\n", "").replace(" ", ""); ad_cat_raw = str(info['cat']).strip()
            if g == "男性":
                if "(限女性)" in ad_item_raw or "(限女性)" in ad_cat_raw or "(女)" in ad_item_raw: continue
            else:
                if "(限男性)" in ad_item_raw or "(限男性)" in ad_cat_raw or "(男)" in ad_item_raw: continue
            is_duplicate = False
            for m_item in main_items_clean:
                if ad_item_clean == m_item: is_duplicate = True; break
            if not is_duplicate: filtered.append((i, info))
        filtered.sort(key=lambda x: x[1]['orig_row'])
        win = tk.Toplevel(self.root); win.title(f"🍭 {self.gt('addon_title')} 🍭"); win.geometry("920x850"); win.configure(bg=AppConfig.COLORS["SEC_BLUE"]); win.grab_set()
        top = tk.Frame(win, bg=AppConfig.COLORS["HEADER_PINK"], pady=10); top.pack(fill=tk.X)
        small_font = ("微软雅黑", 10); small_bold = ("微软雅黑", 10, "bold"); cvs = {i: tk.BooleanVar(value=(i in self.selected_addons)) for i, _ in filtered}
        def save(): self.selected_addons = [idx for idx, v in cvs.items() if v.get()]; self.refresh_ui(); win.destroy()
        tk.Button(top, text=f"💖 {self.gt('save_addon')}", command=save, font=small_bold, bg=AppConfig.COLORS["BTN_GREEN"], fg="white").pack(side=tk.LEFT, padx=20)
        cnt = tk.Frame(win, bg="white", bd=2, relief=tk.SUNKEN); cnt.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        can = tk.Canvas(cnt, bg="white", highlightthickness=0); vsb = ttk.Scrollbar(cnt, orient="vertical", command=can.yview); sf = tk.Frame(can, bg="white")
        sf.bind("<Configure>", lambda e: can.configure(scrollregion=can.bbox("all"))); can.create_window((0,0), window=sf, anchor="nw"); can.configure(yscrollcommand=vsb.set)
        hd = [self.gt("h_select"), self.gt("h_cat"), self.gt("h_item"), self.gt("h_desc"), self.gt("h_price")]; cw = [5, 10, 20, 38, 10]
        for c, t in enumerate(hd): tk.Label(sf, text=t, font=small_bold, relief=tk.RIDGE, bg=AppConfig.COLORS["FOOTER_BUBBLES"], fg="white", width=cw[c], pady=5).grid(row=0, column=c, sticky="nsew")
        cat_counts = []
        if filtered:
            last_cat = filtered[0][1]['cat']; count = 0
            for _, info in filtered:
                if info['cat'] == last_cat: count += 1
                else: cat_counts.append((last_cat, count)); last_cat = info['cat']; count = 1
            cat_counts.append((last_cat, count))
        curr_row = 1
        for cat_name, row_span in cat_counts:
            lbl_cat = tk.Label(sf, text=self.tc(cat_name), relief=tk.RIDGE, bd=1, anchor="nw", justify="left", font=small_font, bg=AppConfig.COLORS["SEC_GREEN"], fg="black", wraplength=100, padx=10, pady=5); lbl_cat.grid(row=curr_row, column=1, rowspan=row_span, sticky="nsew")
            for i in range(row_span):
                idx_in_filtered = curr_row - 1 + i; orig_idx, info = filtered[idx_in_filtered]
                cb_frame = tk.Frame(sf, bg="white", relief=tk.RIDGE, bd=1); cb_frame.grid(row=curr_row + i, column=0, sticky="nsew"); tk.Checkbutton(cb_frame, variable=cvs[orig_idx], bg="white", activebackground="white").pack(expand=True)
                p_text = f"{info['price']} 元" if AppConfig.LANG == "zh" else f"{info['price']} TWD"
                cols = [{"text": self.tc(info['item']), "width": 200, "anchor": "nw", "justify": "left"}, {"text": info['desc'], "width": 380, "anchor": "nw", "justify": "left"}, {"text": p_text, "width": 100, "anchor": "n", "justify": "center"}]
                for c_i, c_cfg in enumerate(cols, 2):
                    lbl = tk.Label(sf, text=c_cfg["text"], relief=tk.RIDGE, bd=1, anchor=c_cfg["anchor"], justify=c_cfg["justify"], font=small_font, bg="white", fg="black", wraplength=c_cfg["width"], padx=10, pady=5); lbl.grid(row=curr_row + i, column=c_i, sticky="nsew"); lbl.bind("<Button-1>", lambda e, x=orig_idx: cvs[x].set(not cvs[x].get()))
            curr_row += row_span
        can.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")
        def _mw(e): can.yview_scroll(int(-1*(e.delta/120)), "units")
        can.bind_all("<MouseWheel>", _mw); win.bind("<Destroy>", lambda e: can.unbind_all("<MouseWheel>"))

    def open_main_package_window(self):
        if self.data_df is None: return
        win = tk.Toplevel(self.root); win.title("🦋 主方案詳情 🦋"); win.geometry("1200x900")
        can = tk.Canvas(win); vsb = ttk.Scrollbar(win, orient="vertical", command=can.yview); hsb = ttk.Scrollbar(win, orient="horizontal", command=can.xview)
        sf = tk.Frame(can); sf.bind("<Configure>", lambda e: can.configure(scrollregion=can.bbox("all"))); can.create_window((0,0), window=sf, anchor="nw"); can.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        lc = ""
        for r in range(len(self.data_df)):
            for c in range(len(self.data_df.columns)):
                v = str(self.data_df.iloc[r, c]); v = "" if v == 'nan' else v
                if c == 0:
                    if not v: v = lc
                    else: lc = v
                bg = AppConfig.COLORS["SEC_BLUE"] if r < 3 else "white"
                tk.Label(sf, text=v, relief=tk.RIDGE, bd=1, bg=bg, padx=5, pady=2, anchor="w", wraplength=350 if c == 2 else 0).grid(row=r, column=c, sticky="nsew")
        can.pack(side="top", fill="both", expand=True); vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")

    def generate_report(self):
        if getattr(self, "_generating", False): return
        is_en = AppConfig.LANG == "en"; display_pkg = self.pkg_var.get(); name = self.name_var.get().strip()
        if not name: messagebox.showwarning("⚠️", "Please enter guest name!" if is_en else "請輸入姓名！"); return
        pkg_k = None
        for k in self.packages.keys():
            if self.tc(k) == display_pkg: pkg_k = k; break
        if not pkg_k: return
        default_name = f"Guest {name} Health Checkup Report.pdf" if is_en else f"貴賓 {name} 健檢項目表.pdf"
        path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=default_name)
        if path:
            self._set_generating(True)
            def worker():
                try:
                    order_map = {}
                    for idx, row in self.items_df.iterrows():
                        item_name = str(row.iloc[1]).strip().replace("\n", "").replace(" ", "")
                        if item_name and item_name != 'nan': order_map[item_name] = idx
                    self.create_pdf(path, name, pkg_k, order_map)
                    self.root.after(0, lambda: messagebox.showinfo("✨", "Report generated successfully!" if is_en else "報表已生成！"))
                except Exception as e:
                    m = str(e); self.root.after(0, lambda x=m: messagebox.showerror("❌", f"PDF Error: {x}"))
                finally: self.root.after(0, lambda: self._set_generating(False))
            threading.Thread(target=worker, daemon=True).start()

    def create_pdf(self, path, name, pkg_key, order_map):
        def get_p(cn): return self.cat_priority.get(self.clean_cat_name(cn), 999)
        is_en = AppConfig.LANG == "en"; gender = self.gender_var.get(); doc = SimpleDocTemplate(path, pagesize=A4); elements = []
        pkg_info = self.packages.get(pkg_key); total_p = self.safe_parse_price(pkg_info["price"])
        styles = getSampleStyleSheet(); title_style = ParagraphStyle('T', fontName='ChineseFont', fontSize=18, alignment=1, spaceAfter=20)
        norm_style = ParagraphStyle('N', fontName='ChineseFont', fontSize=10, leading=14); cat_style = ParagraphStyle('C', fontName='ChineseFont', fontSize=10, leading=14, alignment=1)
        y_text = "2026 Health Checkup Package" if is_en else "2026年健檢專案項目表"; elements.append(Paragraph(f"{y_text} - {name}", title_style))
        elements.append(Paragraph(f"{( 'Guest Name: ' if is_en else '貴賓姓名：')}{name}", norm_style))
        elements.append(Paragraph(f"{( 'Main Package: ' if is_en else '主方案：')}{self.tc(pkg_key)}  |  {( 'Package Price: ' if is_en else '方案價格：')}{pkg_info['price']}", norm_style))
        if self.selected_addons:
            elements.append(Spacer(1, 5)); elements.append(Paragraph("Add-on Items:" if is_en else "單項加選項目：", norm_style))
            addon_table_data = [[Paragraph(f"<b>{( 'Add-on Item' if is_en else '加選項目')}</b>", norm_style), Paragraph(f"<b>{( 'Price' if is_en else '金額')}</b>", norm_style)]]
            sorted_indices = sorted(self.selected_addons, key=lambda i: (get_p(self.add_on_items[i]['cat']), self.add_on_items[i]['item']))
            for i in sorted_indices:
                info = self.add_on_items[i]; p_val = info['price']; total_p += self.safe_parse_price(p_val)
                addon_table_data.append([Paragraph(self.tc(info['item']), norm_style), Paragraph(f"{p_val} {( 'TWD' if is_en else '元')}", norm_style)])
            t_addon = Table(addon_table_data, colWidths=[380, 100]); t_addon.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LEFTPADDING', (0,0), (-1,-1), 10)])); elements.append(t_addon)
        if self.deselected_main_items:
            elements.append(Spacer(1, 5)); elements.append(Paragraph("Removed Items (Deduction):" if is_en else "主方案退選扣除項目：", norm_style))
            for idx in self.deselected_main_items:
                it_name = str(self.data_df.iloc[idx, 1]).strip(); it_cat = ""
                for r in range(idx, 2, -1):
                    c_v = str(self.data_df.iloc[r, 0]).strip()
                    if c_v and c_v.lower() != 'nan': it_cat = c_v; break
                p = self.safe_parse_price(self.addon_price_map.get((it_cat, it_name), "0")); deduction = int(p * 0.8)
                if deduction > 0: total_p -= deduction; elements.append(Paragraph(f" - {self.tc(it_name)} ({( 'Deduct ' if is_en else '扣除 ')}{deduction:,}{( ' TWD' if is_en else ' 元')})", norm_style))
        elements.append(Spacer(1, 10)); elements.append(Paragraph(f"<b>{( 'Estimated Total: ' if is_en else '預估總金額：')}{total_p:,}{( ' TWD' if is_en else ' 元')}</b>", norm_style)); elements.append(Spacer(1, 20))
        pool = []; ex_kw = "限女性" if gender == "男性" else "限男性"; tc = ""
        for idx, row in self.items_df.iterrows():
            c_v = str(self.data_df.iloc[idx, 0]).strip()
            if c_v and c_v.lower() != 'nan': tc = c_v
            val = str(row.iloc[pkg_info["col_idx"]]).strip()
            if val.lower() != 'nan' and val != "":
                if idx in self.deselected_main_items: continue
                item_name = str(row.iloc[1]).strip(); desc = str(row.iloc[2]).strip()
                if '◎' in val and not self.pkg_config_selections.get(item_name, False): continue
                if ex_kw not in item_name and ex_kw not in desc: pool.append({"cat": tc, "item": item_name, "desc": desc, "row_idx": idx})
        for i in self.selected_addons:
            ad = self.add_on_items[i]; ad_item_clean = ad['item'].strip().replace("\n", "").replace(" ", ""); match_idx = 10000 + i
            if ad_item_clean in order_map: match_idx = order_map[ad_item_clean]
            else:
                for main_name, m_idx in order_map.items():
                    if ad_item_clean in main_name or main_name in ad_item_clean: match_idx = m_idx; break
            pool.append({"cat": ad['cat'], "item": ad['item'], "desc": ad['desc'], "row_idx": match_idx})
        pool.sort(key=lambda x: (get_p(x['cat']), x.get('row_idx', 99999)))
        data = [[( 'Category' if is_en else '類別'), ( 'Checkup Item' if is_en else '檢查項目'), ( 'Description' if is_en else '說明')]]
        for it in pool:
            trans_desc = self.tc(it['desc']); final_desc = trans_desc
            if is_en and trans_desc == it['desc'] and any('\u4e00' <= c <= '\u9fff' for c in trans_desc): final_desc = "Consult physician for details"
            data.append([Paragraph(self.tc(it['cat']), cat_style), Paragraph(self.tc(it['item']), norm_style), Paragraph(final_desc, norm_style)])
        t = Table(data, colWidths=[80, 150, 320], repeatRows=1); ts = [('GRID', (0,0), (-1,-1), 0.5, colors.black), ('FONTNAME', (0,0), (-1,-1), 'ChineseFont'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (0,0), (0,-1), 'CENTER'), ('BACKGROUND', (0,0), (-1,0), colors.grey), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke)]
        if pool:
            sr = 1; cur = pool[0]['cat']
            for i in range(1, len(pool)):
                if pool[i]['cat'] != cur:
                    if i > sr: ts.append(('SPAN', (0, sr), (0, i)))
                    sr = i + 1; cur = pool[i]['cat']
            if len(pool) >= sr + 1: ts.append(('SPAN', (0, sr), (0, len(pool))))
        t.setStyle(TableStyle(ts)); elements.append(t); doc.build(elements)

    def generate_excel_report(self):
        if getattr(self, "_generating", False): return
        is_en = AppConfig.LANG == "en"; display_pkg = self.pkg_var.get(); name = self.name_var.get().strip()
        if not name: messagebox.showwarning("⚠️", "Please enter guest name!" if is_en else "請輸入姓名！"); return
        pkg_k = None
        for k in self.packages.keys():
            if self.tc(k) == display_pkg: pkg_k = k; break
        if not pkg_k: return
        default_name = f"Guest {name} Health Checkup.xlsx" if is_en else f"貴賓 {name} 健檢表.xlsx"
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name)
        if path:
            self._set_generating(True)
            def worker():
                try: self.create_excel(path, name, pkg_k); self.root.after(0, lambda: messagebox.showinfo("✨", "Excel report generated!" if is_en else "Excel 報表已生成！"))
                except Exception as e:
                    m = str(e); self.root.after(0, lambda x=m: messagebox.showerror("❌", f"Excel Error: {x}"))
                finally: self.root.after(0, lambda: self._set_generating(False))
            threading.Thread(target=worker, daemon=True).start()

    def _set_generating(self, state):
        self._generating = state; st = "disabled" if state else "normal"; self.gen_btn.config(state=st); self.gen_xlsx_btn.config(state=st)
        if state: self.root.config(cursor="wait"); self.gen_btn.config(text="⌛ Processing..." if AppConfig.LANG=="en" else "⌛ 處理中...")
        else: self.root.config(cursor=""); self.gen_btn.config(text=self.gt("generate"))

    def create_excel(self, path, name, pkg_key):
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        is_en = AppConfig.LANG == "en"; pkg_info = self.packages.get(pkg_key); main_pkg_price = self.safe_parse_price(pkg_info["price"]); total_p = main_pkg_price
        pool = []; gender = self.gender_var.get(); ex_kw = "限女性" if gender == "男性" else "限男性"; tc = ""
        for idx, row in self.items_df.iterrows():
            c_v = str(self.data_df.iloc[idx, 0]).strip()
            if c_v and c_v.lower() != 'nan': tc = c_v
            val = str(row.iloc[pkg_info["col_idx"]]).strip()
            if val.lower() != 'nan' and val != "":
                if idx in self.deselected_main_items: continue
                item_name = str(row.iloc[1]).strip(); desc = str(row.iloc[2]).strip()
                if '◎' in val and not self.pkg_config_selections.get(item_name, False): continue
                if ex_kw not in item_name and ex_kw not in desc: pool.append({"cat": tc, "item": item_name, "desc": desc, "row_idx": idx})
        addon_details = []; addon_total = 0
        for i in self.selected_addons:
            ad = self.add_on_items[i]; p = self.safe_parse_price(ad['price']); addon_total += p; total_p += p; addon_details.append({"name": ad['item'], "price": p})
            pool.append({"cat": ad['cat'], "item": ad['item'], "desc": ad['desc'], "row_idx": 10000+i})
        deduction_details = []; deduction_total = 0
        if self.deselected_main_items:
            for idx in self.deselected_main_items:
                it_name = str(self.data_df.iloc[idx, 1]).strip(); it_cat = ""
                for r in range(idx, 2, -1):
                    c_v = str(self.data_df.iloc[r, 0]).strip()
                    if c_v and c_v.lower() != 'nan': it_cat = c_v; break
                p_raw = self.addon_price_map.get((it_cat, it_name), "0"); p = self.safe_parse_price(p_raw); d_val = int(p * 0.8)
                if d_val > 0: deduction_total += d_val; total_p -= d_val; deduction_details.append({"name": it_name, "val": d_val})
        def get_p(cn): return self.cat_priority.get(self.clean_cat_name(cn), 999)
        pool.sort(key=lambda x: (get_p(x['cat']), x.get('row_idx', 99999))); wb = pd.ExcelWriter(path, engine='openpyxl')
        summary_rows = [[self.gt("ex_name"), name], [self.gt("ex_gender"), self.gt("male") if gender == "男性" else self.gt("female")], [self.gt("ex_pkg_name"), self.tc(pkg_key)], [self.gt("ex_pkg_price"), main_pkg_price], [""]]
        if addon_details:
            summary_rows.append([self.gt("ex_addon_sec")])
            for ad in addon_details: summary_rows.append([f"  + {self.tc(ad['name'])}", ad['price']])
            summary_rows.append([self.gt("ex_addon_sum"), addon_total]); summary_rows.append([""])
        if deduction_details:
            summary_rows.append([self.gt("ex_deduct_sec")])
            for d in deduction_details: summary_rows.append([f"  - {self.tc(d['name'])} (80%)", -d['val']])
            summary_rows.append([self.gt("ex_deduct_sum"), -deduction_total]); summary_rows.append([""])
        summary_rows.append([self.gt("ex_total"), total_p]); summary_rows.append([""])
        all_rows = summary_rows + [[self.gt("h_cat"), self.gt("h_item"), self.gt("h_desc")]]
        for it in pool: all_rows.append([self.tc(it['cat']), self.tc(it['item']), it['desc'] if not is_en else self.tc(it['desc'])])
        df = pd.DataFrame(all_rows); df.to_excel(wb, index=False, header=False, sheet_name="Health Checkup"); ws = wb.sheets["Health Checkup"]
        header_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid"); price_fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid"); title_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"); white_font = Font(color="FFFFFF", bold=True, size=12); border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 40; ws.column_dimensions['C'].width = 80
        for r in range(1, len(summary_rows) + 1):
            cell_a = ws.cell(row=r, column=1); cell_b = ws.cell(row=r, column=2); val_a = cell_a.value
            if val_a:
                cell_a.font = Font(bold=True)
                if "總金額" in str(val_a) or "預估" in str(val_a) or "Total" in str(val_a): cell_a.font = Font(bold=True, size=14, color="FF1493"); cell_a.fill = price_fill; cell_b.font = Font(bold=True, size=14, color="FF1493")
                elif any(k in str(val_a) for k in ["價格", "總計", "項目", "Price", "Total"]): cell_a.fill = price_fill
                else: cell_a.fill = title_fill
            if isinstance(cell_b.value, (int, float)): cell_b.number_format = '#,##0'; cell_b.alignment = Alignment(horizontal="right")
        detail_header_row = len(summary_rows) + 1
        for c in range(1, 4): cell = ws.cell(row=detail_header_row, column=c); cell.fill = header_fill; cell.font = white_font; cell.alignment = Alignment(horizontal="center"); cell.border = border
        start_row = detail_header_row + 1
        if pool:
            cur_cat = pool[0]['cat']; merge_start = start_row
            for i, it in enumerate(pool):
                curr_r = start_row + i
                for c in range(1, 4): cell = ws.cell(row=curr_r, column=c); cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = border
                if it['cat'] != cur_cat:
                    if curr_r - 1 > merge_start: ws.merge_cells(start_row=merge_start, start_column=1, end_row=curr_r-1, end_column=1)
                    merge_start = curr_r; cur_cat = it['cat']
            if start_row + len(pool) - 1 > merge_start: ws.merge_cells(start_row=merge_start, start_column=1, end_row=start_row + len(pool) - 1, end_column=1)
        wb.close()

if __name__ == "__main__":
    root = tk.Tk(); app = HealthCheckupApp(root); root.mainloop()
